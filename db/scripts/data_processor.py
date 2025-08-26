#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
IoTデータ処理システム
Excel/CSVファイルを監視し、データを抽出してPostgreSQLに格納
"""

import os
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import List, Dict, Tuple, Optional
import logging

import pandas as pd
import numpy as np
import psycopg2
from psycopg2.extras import execute_values
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler, FileCreatedEvent
import openpyxl
from pydantic import BaseModel, validator, Field
from loguru import logger
import traceback

# ログ設定
logger.remove()  # デフォルトのハンドラを削除
logger.add(sys.stderr, level="INFO")
logger.add("logs/iot_processor.log", rotation="1 day",
           retention="30 days", level="DEBUG")

# ===============================================
# 設定クラス
# ===============================================


class Config(BaseModel):
    """アプリケーション設定"""
    # データベース接続
    db_host: str = "localhost"
    db_port: int = 5432
    db_name: str = "iot_monitor"
    db_user: str = "postgres"
    db_password: str = "postgres"

    # ファイル監視
    watch_directory: str = "./data/incoming"
    processed_directory: str = "./data/processed"
    error_directory: str = "./data/error"

    # ファイル処理
    file_extensions: List[str] = [".xlsx", ".csv"]
    batch_size: int = 10000  # バッチ挿入サイズ

    # Excel設定
    tag_row: int = 36  # タグコードの行（1ベース）
    data_start_row: int = 40  # データ開始行（1ベース）
    date_column: str = "A"  # 日付列
    time_column: str = "B"  # 時間列
    tag_start_column: str = "D"  # タグ開始列
    tag_column_interval: int = 2  # タグ列の間隔

    class Config:
        env_file = ".env"

# ===============================================
# データモデル
# ===============================================


class MeasurementData(BaseModel):
    """測定データモデル"""
    timestamp: datetime
    tag_id: int
    value: float

    @validator('value')
    def validate_value(cls, v):
        if pd.isna(v) or np.isnan(v):
            raise ValueError("Value cannot be NaN")
        return v

# ===============================================
# データベース管理クラス
# ===============================================


class DatabaseManager:
    """PostgreSQL接続管理"""

    def __init__(self, config: Config):
        self.config = config
        self.conn = None
        self.tag_cache = {}  # tag_code -> tag_id のキャッシュ

    def connect(self):
        """データベース接続"""
        try:
            self.conn = psycopg2.connect(
                host=self.config.db_host,
                port=self.config.db_port,
                database=self.config.db_name,
                user=self.config.db_user,
                password=self.config.db_password
            )
            self.conn.autocommit = False
            logger.info("データベースに接続しました")
            self._load_tag_cache()
        except Exception as e:
            logger.error(f"データベース接続エラー: {e}")
            raise

    def disconnect(self):
        """データベース切断"""
        if self.conn:
            self.conn.close()
            logger.info("データベース接続を切断しました")

    def _load_tag_cache(self):
        """タグ情報をキャッシュに読み込み"""
        with self.conn.cursor() as cur:
            cur.execute("""
                SELECT tag_id, tag_code, min_value, max_value 
                FROM tags 
                WHERE tag_code IS NOT NULL
            """)
            for row in cur.fetchall():
                self.tag_cache[row[1]] = {
                    'tag_id': row[0],
                    'min_value': row[2],
                    'max_value': row[3]
                }
        logger.info(f"{len(self.tag_cache)}個のタグ情報をキャッシュしました")

    def validate_tag_code(self, tag_code: str) -> Optional[Dict]:
        """タグコードの検証"""
        return self.tag_cache.get(tag_code)

    def insert_measurements(self, measurements: List[MeasurementData]):
        """測定データの一括挿入"""
        if not measurements:
            return

        try:
            with self.conn.cursor() as cur:
                # データを準備
                data = [
                    (m.timestamp, m.tag_id, m.value)
                    for m in measurements
                ]

                # 一括挿入（高速）
                execute_values(
                    cur,
                    """
                    INSERT INTO measurements (timestamp, tag_id, value)
                    VALUES %s
                    ON CONFLICT (timestamp, tag_id) DO UPDATE
                    SET value = EXCLUDED.value
                    """,
                    data,
                    template="(%s, %s, %s)"
                )

            self.conn.commit()
            logger.info(f"{len(measurements)}件のデータを挿入しました")

        except Exception as e:
            self.conn.rollback()
            logger.error(f"データ挿入エラー: {e}")
            raise

    def refresh_materialized_views(self):
        """マテリアライズドビューの更新"""
        try:
            with self.conn.cursor() as cur:
                # 優先度順に更新
                views_to_refresh = [
                    'mv_power_1min',
                    'mv_temp_1min',
                    'mv_humid_5min',
                    'mv_temp_5min',
                    'mv_integrated_power_30min'
                ]

                for view in views_to_refresh:
                    cur.execute(
                        f"REFRESH MATERIALIZED VIEW {view}")
                    logger.info(f"更新完了: {view}")

            self.conn.commit()
        except Exception as e:
            logger.error(f"MV更新エラー: {e}")
            self.conn.rollback()

# ===============================================
# Excel/CSVファイル処理クラス
# ===============================================


class DataFileProcessor:
    """データファイル処理"""

    def __init__(self, config: Config, db_manager: DatabaseManager):
        self.config = config
        self.db_manager = db_manager

    def process_file(self, file_path: str) -> bool:
        """ファイル処理のメインメソッド"""
        logger.info(f"ファイル処理開始: {file_path}")

        try:
            # ファイル拡張子で処理を分岐
            if file_path.endswith('.xlsx'):
                data = self._read_excel_file(file_path)
            elif file_path.endswith('.csv'):
                data = self._read_csv_file(file_path)
            else:
                raise ValueError(f"サポートされていないファイル形式: {file_path}")

            # データをデータベースに保存
            self._save_to_database(data)

            # データ処理後にMVを更新
            self.db_manager.refresh_materialized_views()

            # 処理済みフォルダに移動
            self._move_processed_file(file_path)

            logger.info(f"ファイル処理完了: {file_path}")
            return True

        except Exception as e:
            logger.error(f"ファイル処理エラー: {file_path} - {e}")
            logger.error(traceback.format_exc())
            self._move_error_file(file_path)
            return False

    def _read_excel_file(self, file_path: str) -> List[MeasurementData]:
        """Excelファイルの読み込み"""
        measurements = []

        # openpyxlで読み込み（メモリ効率的）
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        # タグコードの取得（36行目）
        tag_codes = {}
        col_idx = openpyxl.utils.column_index_from_string(
            self.config.tag_start_column)

        # 最終列を動的に取得
        max_col = ws.max_column
        logger.debug(f"最大列数: {max_col}")

        # タグコードを収集
        for col in range(col_idx, max_col + 1, self.config.tag_column_interval):
            cell_value = ws.cell(row=self.config.tag_row, column=col).value
            if cell_value:
                tag_info = self.db_manager.validate_tag_code(str(cell_value))
                if tag_info:
                    tag_codes[col] = tag_info
                    logger.debug(f"タグ発見: 列{col} = {cell_value}")
                else:
                    logger.warning(f"未登録のタグコード: {cell_value}")

        logger.info(f"{len(tag_codes)}個の有効なタグを発見")

        # データ行の処理（40行目以降）
        row_count = 0
        for row in ws.iter_rows(min_row=self.config.data_start_row, values_only=True):
            # 日付と時間の取得
            date_value = row[0]  # A列
            time_value = row[1]  # B列

            # 日付/時間がない場合はスキップ
            if not date_value or not time_value:
                continue

            # タイムスタンプの作成
            try:
                timestamp = self._create_timestamp(date_value, time_value)
            except Exception as e:
                logger.warning(
                    f"タイムスタンプ作成エラー（行{row_count + self.config.data_start_row}）: {e}")
                continue

            # 各タグの値を処理
            for col_idx, tag_info in tag_codes.items():
                try:
                    # 列インデックスを0ベースに変換
                    value_idx = col_idx - 1
                    if value_idx < len(row):
                        value = row[value_idx]

                        # 値の検証
                        if value is not None and not pd.isna(value):
                            # 数値に変換
                            value = float(value)

                            # 範囲チェック
                            min_val = tag_info.get('min_value')
                            max_val = tag_info.get('max_value')

                            if min_val is not None and value < min_val:
                                logger.warning(
                                    f"値が最小値未満: tag_id={tag_info['tag_id']}, value={value}, min={min_val}")
                                continue
                            if max_val is not None and value > max_val:
                                logger.warning(
                                    f"値が最大値超過: tag_id={tag_info['tag_id']}, value={value}, max={max_val}")
                                continue

                            # 測定データを作成
                            measurement = MeasurementData(
                                timestamp=timestamp,
                                tag_id=tag_info['tag_id'],
                                value=value
                            )
                            measurements.append(measurement)

                except Exception as e:
                    logger.debug(
                        f"値処理エラー（行{row_count + self.config.data_start_row}, 列{col_idx}）: {e}")

            row_count += 1

            # バッチ処理
            if len(measurements) >= self.config.batch_size:
                self.db_manager.insert_measurements(measurements)
                measurements = []

        # 残りのデータを挿入
        if measurements:
            self.db_manager.insert_measurements(measurements)

        wb.close()
        logger.info(f"{row_count}行を処理しました")

        return measurements

    def _read_csv_file(self, file_path: str) -> List[MeasurementData]:
        """CSVファイルの読み込み（将来の実装用）"""
        # CSVの場合も基本的にはExcelと同じ処理
        # pandasを使用して効率的に読み込む
        raise NotImplementedError("CSV処理は未実装です")

    def _create_timestamp(self, date_value, time_value) -> datetime:
        """日付と時間からタイムスタンプを作成"""
        # Excelの日付シリアル値の場合
        if isinstance(date_value, (int, float)):
            # Excel基準日（1900-01-01）からの日数
            base_date = datetime(1899, 12, 30)  # Excelのバグ対応
            date_obj = base_date + pd.Timedelta(days=int(date_value))
        elif isinstance(date_value, datetime):
            date_obj = date_value
        else:
            # 文字列の場合
            date_obj = pd.to_datetime(str(date_value))

        # 時間の処理
        if isinstance(time_value, (int, float)):
            # Excelの時間（0-1の小数）
            hours = int(time_value * 24)
            minutes = int((time_value * 24 - hours) * 60)
            seconds = int(((time_value * 24 - hours) * 60 - minutes) * 60)
        elif isinstance(time_value, datetime):
            hours = time_value.hour
            minutes = time_value.minute
            seconds = time_value.second
        else:
            # 文字列の場合
            time_parts = str(time_value).split(':')
            hours = int(time_parts[0])
            minutes = int(time_parts[1]) if len(time_parts) > 1 else 0
            seconds = int(time_parts[2]) if len(time_parts) > 2 else 0

        # タイムスタンプの作成（タイムゾーン付き）
        timestamp = date_obj.replace(
            hour=hours,
            minute=minutes,
            second=seconds,
            tzinfo=timezone.utc
        )

        return timestamp

    def _save_to_database(self, measurements: List[MeasurementData]):
        """データベースへの保存（既にinsert_measurementsで処理済み）"""
        pass

    def _move_processed_file(self, file_path: str):
        """処理済みファイルの移動"""
        processed_path = Path(self.config.processed_directory)
        processed_path.mkdir(parents=True, exist_ok=True)

        file_name = Path(file_path).name
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        new_name = f"{timestamp}_{file_name}"
        new_path = processed_path / new_name

        Path(file_path).rename(new_path)
        logger.info(f"ファイルを移動: {file_path} -> {new_path}")

    def _move_error_file(self, file_path: str):
        """エラーファイルの移動"""
        error_path = Path(self.config.error_directory)
        error_path.mkdir(parents=True, exist_ok=True)

        file_name = Path(file_path).name
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        new_name = f"ERROR_{timestamp}_{file_name}"
        new_path = error_path / new_name

        Path(file_path).rename(new_path)
        logger.warning(f"エラーファイルを移動: {file_path} -> {new_path}")

# ===============================================
# ファイル監視クラス
# ===============================================


class FileWatcher(FileSystemEventHandler):
    """ファイルシステム監視"""

    def __init__(self, processor: DataFileProcessor, config: Config):
        self.processor = processor
        self.config = config
        self.processing = set()  # 処理中のファイル

    def on_created(self, event: FileCreatedEvent):
        """ファイル作成イベント"""
        if event.is_directory:
            return

        file_path = event.src_path

        # サポートされている拡張子かチェック
        if not any(file_path.endswith(ext) for ext in self.config.file_extensions):
            return

        # ファイルが完全に書き込まれるまで待機
        time.sleep(2)

        # 既に処理中でないかチェック
        if file_path in self.processing:
            return

        try:
            self.processing.add(file_path)
            logger.info(f"新規ファイル検出: {file_path}")

            # ファイル処理
            self.processor.process_file(file_path)

        finally:
            self.processing.discard(file_path)

# ===============================================
# メインアプリケーション
# ===============================================


class IoTDataProcessor:
    """IoTデータ処理アプリケーション"""

    def __init__(self, config: Config):
        self.config = config
        self.db_manager = DatabaseManager(config)
        self.processor = DataFileProcessor(config, self.db_manager)
        self.file_watcher = FileWatcher(self.processor, config)
        self.observer = Observer()

    def start(self):
        """アプリケーション開始"""
        logger.info("IoTデータ処理システムを開始します")

        # ディレクトリ作成
        for directory in [self.config.watch_directory,
                          self.config.processed_directory,
                          self.config.error_directory]:
            Path(directory).mkdir(parents=True, exist_ok=True)

        # データベース接続
        self.db_manager.connect()

        # ファイル監視開始
        self.observer.schedule(
            self.file_watcher,
            self.config.watch_directory,
            recursive=False
        )
        self.observer.start()

        logger.info(f"ファイル監視を開始: {self.config.watch_directory}")

        try:
            # 既存ファイルの処理
            self._process_existing_files()

            # 監視を継続
            while True:
                time.sleep(1)

        except KeyboardInterrupt:
            logger.info("終了シグナルを受信しました")
        finally:
            self.stop()

    def stop(self):
        """アプリケーション停止"""
        logger.info("IoTデータ処理システムを停止します")

        self.observer.stop()
        self.observer.join()
        self.db_manager.disconnect()

    def _process_existing_files(self):
        """既存ファイルの処理"""
        watch_path = Path(self.config.watch_directory)

        for file_path in watch_path.iterdir():
            if file_path.is_file() and any(str(file_path).endswith(ext)
                                           for ext in self.config.file_extensions):
                logger.info(f"既存ファイル発見: {file_path}")
                self.processor.process_file(str(file_path))

# ===============================================
# エントリーポイント
# ===============================================


def main():
    """メイン関数"""
    # 設定読み込み
    config = Config()

    # アプリケーション起動
    app = IoTDataProcessor(config)
    app.start()


if __name__ == "__main__":
    main()
