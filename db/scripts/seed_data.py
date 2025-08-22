import pandas as pd
import psycopg2
from openpyxl import load_workbook

EXCEL_PATH = "data/tbl/テーブル設計.xlsm"
SHEET_NAME = "テーブル"


def read_named_table(path, sheet, table_name):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]
    for tbl in ws._tables.values():
        if tbl.name == table_name:
            ref = tbl.ref
            start_cell, end_cell = ref.split(":")
            start_row = ws[start_cell].row - 1
            start_col = ws[start_cell].column - 1
            end_row = ws[end_cell].row
            end_col = ws[end_cell].column
            df = pd.read_excel(
                path,
                sheet_name=sheet,
                header=start_row,
                usecols=range(start_col, end_col)
            )
            return df
    raise ValueError(f"Table '{table_name}' not found in sheet '{sheet}'.")


# 各テーブルのデータ取得＆空行除去
df_buildings = read_named_table(EXCEL_PATH, SHEET_NAME, "tbl_buildings")
df_buildings = df_buildings[df_buildings["building_name"].notnull()].copy()

df_locations = read_named_table(EXCEL_PATH, SHEET_NAME, "tbl_locations")
df_locations = df_locations[df_locations["location_name"].notnull()].copy()

df_measure_types = read_named_table(
    EXCEL_PATH, SHEET_NAME, "tbl_measure_types")
df_measure_types = df_measure_types[df_measure_types["measure_type_name"].notnull(
)].copy()

df_tags = read_named_table(EXCEL_PATH, SHEET_NAME, "tbl_tags")
df_tags = df_tags[df_tags["tag_code"].notnull()].copy()

# カラム名補正
for col in df_locations.columns:
    if "building_id" in col:
        df_locations = df_locations.rename(columns={col: "building_id"})
    if "location_id" in col:
        df_locations = df_locations.rename(columns={col: "location_id"})

for col in df_measure_types.columns:
    if "measure_type_id" in col:
        df_measure_types = df_measure_types.rename(
            columns={col: "measure_type_id"})

for col in df_tags.columns:
    if "tag_id" in col:
        df_tags = df_tags.rename(columns={col: "tag_id"})
    if "building_id" in col:
        df_tags = df_tags.rename(columns={col: "building_id"})
    if "location_id" in col:
        df_tags = df_tags.rename(columns={col: "location_id"})
    if "measure_type_id" in col:
        df_tags = df_tags.rename(columns={col: "measure_type_id"})

print("buildings (preview):")
print(df_buildings)
print("locations (preview):")
print(df_locations)
print("measure_types (preview):")
print(df_measure_types)
print("tags (preview):")
print(df_tags)

conn_dict = {
    "host": "localhost", "port": 5432,
    "dbname": "iot_monitor",
    "user": "postgres",
    "password": "postgres"
}

with psycopg2.connect(**conn_dict) as conn:
    with conn.cursor() as cur:
        # 順番: 子テーブル→親テーブルの順にdelete
        cur.execute("DELETE FROM tags;")
        cur.execute("DELETE FROM locations;")
        cur.execute("DELETE FROM buildings;")
        cur.execute("DELETE FROM measure_types;")
        # measure_types
        for _, row in df_measure_types.iterrows():
            cur.execute(
                "INSERT INTO measure_types (measure_type_id, measure_type_name, unit, created_at, updated_at) VALUES (%s, %s, %s, DEFAULT, DEFAULT);",
                (int(row["measure_type_id"]),
                 row["measure_type_name"], row["unit"])
            )
        # buildings
        for _, row in df_buildings.iterrows():
            cur.execute(
                "INSERT INTO buildings (building_id, building_name, created_at, updated_at) VALUES (%s, %s, DEFAULT, DEFAULT);",
                (int(row["building_id"]), row["building_name"])
            )
        # locations
        for _, row in df_locations.iterrows():
            cur.execute(
                "INSERT INTO locations (location_id, building_id, location_name, floor, is_active, created_at, updated_at) VALUES (%s, %s, %s, %s, %s, DEFAULT, DEFAULT);",
                (int(row["location_id"]), int(row["building_id"]),
                 row["location_name"], row["floor"], row["is_active"])
            )
        # tags
        for _, row in df_tags.iterrows():
            cur.execute(
                """
                INSERT INTO tags (
                    tag_id, building_id, location_id, measure_type_id,
                    tag_code, min_value, max_value, created_at, updated_at
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, DEFAULT, DEFAULT);
                """,
                (
                    int(row["tag_id"]),
                    int(row["building_id"]),
                    int(row["location_id"]),
                    int(row["measure_type_id"]),
                    row["tag_code"],
                    row["min_value"],
                    row["max_value"]
                )
            )
        conn.commit()

print("tbl_buildings, tbl_locations, tbl_measure_types, tbl_tagsのデータをDB登録しました。")
